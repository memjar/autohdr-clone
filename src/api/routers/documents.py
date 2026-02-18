"""
Document Intelligence Router — Upload, parse, embed, and query documents.
Supports PDF, DOCX, XLSX, CSV, and TXT files.
Includes Google Drive ingestion for shared/public links.
"""

import os
import re
import tempfile
import logging

import httpx
from fastapi import APIRouter, UploadFile, File, HTTPException
from pydantic import BaseModel
from typing import Optional

import imi_rag

logger = logging.getLogger("imi_documents")

router = APIRouter(prefix="/klaus/imi", tags=["imi"])


# --- Text extraction helpers ---

def _extract_pdf(path: str) -> str:
    from pdfminer.high_level import extract_text
    return extract_text(path)


def _extract_docx(path: str) -> str:
    from docx import Document
    doc = Document(path)
    return "\n".join(p.text for p in doc.paragraphs)


def _extract_xlsx(path: str) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    lines = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            lines.append(" | ".join(cells))
    wb.close()
    return "\n".join(lines)


def _extract_text(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        return f.read()


EXTRACTORS = {
    ".pdf": _extract_pdf,
    ".docx": _extract_docx,
    ".xlsx": _extract_xlsx,
    ".csv": _extract_text,
    ".txt": _extract_text,
}

ALLOWED_EXTENSIONS = set(EXTRACTORS.keys())


def _chunk_text(text: str, chunk_size: int = 500, overlap: int = 100) -> list[str]:
    """Split text into overlapping chunks."""
    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunks.append(text[start:end])
        start += chunk_size - overlap
    return chunks


# --- Endpoints ---

@router.post("/upload-document")
async def upload_document(file: UploadFile = File(...)):
    """Upload and ingest a document into the vector store."""
    filename = file.filename or "unknown"
    ext = os.path.splitext(filename)[1].lower()

    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type '{ext}'. Supported: {', '.join(sorted(ALLOWED_EXTENSIONS))}"
        )

    # Write to temp file, extract, then delete
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=ext)
    try:
        content = await file.read()
        with os.fdopen(tmp_fd, "wb") as tmp_f:
            tmp_f.write(content)

        extractor = EXTRACTORS[ext]
        text = extractor(tmp_path)
    except Exception as e:
        logger.error(f"Failed to extract text from {filename}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to process file: {e}")
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)

    if not text.strip():
        raise HTTPException(status_code=400, detail="No text content could be extracted from file.")

    # Use row-based chunking for CSV, standard for others
    if ext == ".csv":
        chunks = _chunk_csv_by_rows(text, rows_per_chunk=50)
    else:
        chunks = _chunk_text(text)
    result = imi_rag.ingest_document(filename, chunks)

    return {
        "status": "success",
        "filename": filename,
        "chunks": result["documents"],
        "dataset": result["dataset"],
    }


@router.get("/documents")
async def list_documents():
    """List all uploaded documents in the vector store."""
    store = imi_rag._load_store()
    datasets = {}
    for doc in store.get("documents", []):
        meta = doc.get("metadata", {})
        if meta.get("category") == "uploaded":
            ds = meta.get("dataset", "unknown")
            datasets[ds] = datasets.get(ds, 0) + 1

    return {
        "documents": [
            {"dataset": ds, "chunks": count}
            for ds, count in sorted(datasets.items())
        ]
    }


class QueryRequest(BaseModel):
    query: str
    dataset: Optional[str] = None


@router.post("/query-document")
async def query_document(req: QueryRequest):
    """Query uploaded documents with optional dataset filter."""
    # Use imi_rag.query with category filter for uploaded docs
    results = imi_rag.query(req.query, n=10, category="uploaded")

    # Filter by specific dataset if provided
    if req.dataset:
        results = [r for r in results if r.get("dataset") == req.dataset]

    return {"query": req.query, "dataset": req.dataset, "results": results}


# --- Google Drive ingestion ---

def _extract_gdrive_file_id(url: str) -> str | None:
    """Extract file ID from various Google Drive URL formats."""
    patterns = [
        r"/file/d/([a-zA-Z0-9_-]+)",       # /file/d/FILE_ID/
        r"id=([a-zA-Z0-9_-]+)",             # ?id=FILE_ID
        r"/d/([a-zA-Z0-9_-]+)",             # /d/FILE_ID/
        r"open\?id=([a-zA-Z0-9_-]+)",       # open?id=FILE_ID
        r"/spreadsheets/d/([a-zA-Z0-9_-]+)", # Google Sheets
    ]
    for p in patterns:
        m = re.search(p, url)
        if m:
            return m.group(1)
    return None


def _guess_extension_from_headers(headers: dict, url: str) -> str:
    """Guess file extension from content-type or URL."""
    ct = headers.get("content-type", "")
    if "spreadsheet" in ct or "excel" in ct:
        return ".xlsx"
    if "csv" in ct or "comma" in ct:
        return ".csv"
    if "pdf" in ct:
        return ".pdf"
    if "word" in ct or "msword" in ct or "officedocument.wordprocessing" in ct:
        return ".docx"
    if "text/plain" in ct:
        return ".txt"
    # Fallback: check URL
    if "export=csv" in url or "gviz/tq?tqx=out:csv" in url:
        return ".csv"
    return ".csv"  # Default to CSV for Google Sheets


def _chunk_csv_by_rows(text: str, rows_per_chunk: int = 50) -> list[str]:
    """Chunk CSV by rows, keeping the header with each chunk for context."""
    lines = text.strip().split("\n")
    if len(lines) <= 1:
        return [text]
    header = lines[0]
    data_lines = lines[1:]
    chunks = []
    for i in range(0, len(data_lines), rows_per_chunk):
        batch = data_lines[i:i + rows_per_chunk]
        chunks.append(header + "\n" + "\n".join(batch))
    return chunks


class GDriveRequest(BaseModel):
    url: str
    name: Optional[str] = None


@router.post("/ingest-gdrive")
async def ingest_gdrive(req: GDriveRequest):
    """
    Ingest a file from Google Drive (public or shared link).
    Supports Sheets (exported as CSV), PDFs, DOCX, XLSX, CSV, TXT.
    """
    file_id = _extract_gdrive_file_id(req.url)
    if not file_id:
        raise HTTPException(status_code=400, detail="Could not extract file ID from URL. Paste a Google Drive share link.")

    # Try Google Sheets CSV export first, then direct download
    download_urls = [
        f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=csv",
        f"https://drive.google.com/uc?export=download&id={file_id}",
        f"https://drive.usercontent.google.com/download?id={file_id}&export=download",
    ]

    content = None
    used_url = ""
    headers = {}

    async with httpx.AsyncClient(follow_redirects=True, timeout=120.0) as client:
        for dl_url in download_urls:
            try:
                resp = await client.get(dl_url)
                if resp.status_code == 200 and len(resp.content) > 100:
                    # Check for Google's "virus scan" confirmation page
                    if b"confirm=" in resp.content and b"Google Drive" in resp.content:
                        # Extract confirm token and retry
                        confirm_match = re.search(r"confirm=([a-zA-Z0-9_-]+)", resp.text)
                        if confirm_match:
                            confirm_url = f"{dl_url}&confirm={confirm_match.group(1)}"
                            resp = await client.get(confirm_url)
                    if resp.status_code == 200 and len(resp.content) > 100:
                        content = resp.content
                        used_url = dl_url
                        headers = dict(resp.headers)
                        break
            except httpx.HTTPError:
                continue

    if content is None:
        raise HTTPException(
            status_code=400,
            detail="Could not download file. Make sure the link is set to 'Anyone with the link can view'."
        )

    ext = _guess_extension_from_headers(headers, used_url)
    filename = req.name or f"gdrive_{file_id}{ext}"
    if not filename.endswith(ext):
        filename += ext

    # Write to temp, extract, chunk, ingest
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=ext)
    try:
        with os.fdopen(tmp_fd, "wb") as tmp_f:
            tmp_f.write(content)

        extractor = EXTRACTORS.get(ext, _extract_text)
        text = extractor(tmp_path)
    except Exception as e:
        logger.error(f"Failed to extract Google Drive file {filename}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to process file: {e}")
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)

    if not text.strip():
        raise HTTPException(status_code=400, detail="No text content could be extracted from file.")

    # Use row-based chunking for CSV (better for large tabular data)
    if ext == ".csv":
        chunks = _chunk_csv_by_rows(text, rows_per_chunk=50)
    else:
        chunks = _chunk_text(text)

    result = imi_rag.ingest_document(filename, chunks)

    row_count = text.count("\n") if ext == ".csv" else None
    size_kb = len(content) / 1024

    return {
        "status": "success",
        "filename": filename,
        "source": "google_drive",
        "file_id": file_id,
        "size_kb": round(size_kb, 1),
        "rows": row_count,
        "chunks": result["documents"],
        "dataset": result["dataset"],
        "message": f"Ingested {result['documents']} chunks from {filename}. Klaus can now answer questions about this data.",
    }
