"""
AutoHDR Clone - FastAPI Backend
================================

REST API for real estate photo editing with full RAW file support.

Supports: ARW (Sony), CR2/CR3 (Canon), NEF (Nikon), DNG, RAF (Fuji), etc.

Run:
    ./start-backend.sh

Endpoints:
    POST /process   - Main processing (HDR merge or twilight)
    GET  /health    - Health check
    GET  /test      - Test processing with generated image
"""

import io
import os
import time
import traceback
from datetime import datetime
from pathlib import Path
from typing import List, Optional

import cv2
import numpy as np
from fastapi import FastAPI, File, UploadFile, HTTPException, Form, Query, Request
from fastapi.responses import StreamingResponse, Response, JSONResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

# ============================================
# DEPENDENCY CHECKS
# ============================================

# RAW file support
try:
    import rawpy
    HAS_RAWPY = True
    RAWPY_VERSION = rawpy.__version__ if hasattr(rawpy, '__version__') else 'unknown'
except ImportError:
    HAS_RAWPY = False
    RAWPY_VERSION = None
    print("⚠️  Warning: rawpy not installed. RAW file support disabled.")
    print("   Install with: pip install rawpy")

# OpenCV version
CV2_VERSION = cv2.__version__

# Import our processor
import sys
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

try:
    from src.core.processor import AutoHDRProcessor, ProcessingSettings, PROCESSOR_VERSION
    HAS_PROCESSOR = True
except ImportError as e:
    HAS_PROCESSOR = False
    PROCESSOR_VERSION = None
    print(f"⚠️  Warning: Could not import processor: {e}")

# AI-enhanced processor (optional)
try:
    from src.core.processor_ai import AIEnhancedProcessor, AIProcessingSettings
    HAS_AI_PROCESSOR = True
except ImportError as e:
    HAS_AI_PROCESSOR = False
    print(f"ℹ️  AI processor not available (install ai-requirements.txt for 90% quality)")

# Pro Processor v3.0 - THE golden implementation (95%+ AutoHDR quality)
try:
    from src.core.processor_v3 import AutoHDRProProcessor, ProSettings, PROCESSOR_VERSION as PRO_VERSION
    HAS_PRO_PROCESSOR = True
    print(f"✓ Pro Processor v{PRO_VERSION} loaded - 95%+ AutoHDR quality")
except ImportError as e:
    HAS_PRO_PROCESSOR = False
    PRO_VERSION = None
    print(f"ℹ️  Pro Processor not available: {e}")

# Optional modules
HAS_HDR_MERGER = False
HAS_TWILIGHT = False

# ============================================
# APP SETUP
# ============================================

app = FastAPI(
    title="AutoHDR Clone API",
    description="Open-source AI real estate photo editing with RAW support",
    version="1.0.0",
    docs_url="/docs",
    redoc_url="/redoc",
)

# CORS for web frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "https://autohdr-clone.vercel.app",
        "https://*.vercel.app",
        "*"  # Allow all for development
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Survey data store (used by IMI endpoints — declared early so all routes can access)
_SURVEY_STORE: dict = {}  # survey_id -> {"structured_data": dict, "raw_text": str, ...}

# ============================================
# STARTUP EVENT
# ============================================

@app.on_event("startup")
async def startup_event():
    """Print status on startup."""
    print("")
    print("╔══════════════════════════════════════════════════════════════╗")
    print("║           AutoHDR Clone - Backend Server                     ║")
    print("╠══════════════════════════════════════════════════════════════╣")
    print(f"║  OpenCV:     {CV2_VERSION:<10} ✓                                 ║")
    if HAS_RAWPY:
        print(f"║  rawpy:      {RAWPY_VERSION:<10} ✓  (ARW/CR2/NEF support)       ║")
    else:
        print("║  rawpy:      NOT INSTALLED ✗  (RAW files disabled)        ║")
    if HAS_PROCESSOR:
        version_str = PROCESSOR_VERSION if PROCESSOR_VERSION else "Ready"
        print(f"║  Processor:  {version_str:<10} ✓                                   ║")
    else:
        print("║  Processor:  NOT FOUND  ✗                                   ║")
    print("╠══════════════════════════════════════════════════════════════╣")
    print("║  Endpoints:                                                   ║")
    print("║    POST /process   - Process images (HDR or twilight)         ║")
    print("║    GET  /test      - Test processing pipeline                 ║")
    print("║    GET  /health    - Health check                             ║")
    print("║    GET  /docs      - API documentation                        ║")
    print("╚══════════════════════════════════════════════════════════════╝")
    print("")


# ============================================
# RAW FILE EXTENSIONS
# ============================================

RAW_EXTENSIONS = {
    '.arw', '.srf', '.sr2',  # Sony
    '.cr2', '.cr3', '.crw',  # Canon
    '.nef', '.nrw',          # Nikon
    '.dng',                  # Adobe
    '.orf',                  # Olympus
    '.rw2',                  # Panasonic
    '.pef', '.ptx',          # Pentax
    '.raf',                  # Fujifilm
    '.erf',                  # Epson
    '.mrw',                  # Minolta
    '.3fr', '.fff',          # Hasselblad
    '.iiq',                  # Phase One
    '.rwl',                  # Leica
    '.srw',                  # Samsung
    '.x3f',                  # Sigma
    '.raw',                  # Generic
}


# ============================================
# MODELS
# ============================================

class HDRMergeRequest(BaseModel):
    method: str = "mertens"
    tone_map: str = "reinhard"
    align: bool = True


class TwilightRequest(BaseModel):
    sky_intensity: float = 0.9
    window_glow: float = 0.8


class ProcessingResponse(BaseModel):
    success: bool
    job_id: str
    message: str


# ============================================
# HELPERS
# ============================================

def read_image_from_upload(file: UploadFile) -> np.ndarray:
    """Read uploaded file into OpenCV image, including RAW formats."""
    contents = file.file.read()
    filename = file.filename or "image.jpg"
    ext = Path(filename).suffix.lower()

    # Handle RAW files
    if ext in RAW_EXTENSIONS:
        if not HAS_RAWPY:
            raise HTTPException(400, f"RAW file support not available. Install rawpy.")
        try:
            with rawpy.imread(io.BytesIO(contents)) as raw:
                rgb = raw.postprocess(
                    use_camera_wb=True,
                    half_size=False,
                    no_auto_bright=False,
                    output_bps=8
                )
                # Convert RGB to BGR for OpenCV
                return cv2.cvtColor(rgb, cv2.COLOR_RGB2BGR)
        except Exception as e:
            raise HTTPException(400, f"Could not decode RAW file {filename}: {str(e)}")

    # Standard image formats
    nparr = np.frombuffer(contents, np.uint8)
    image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    if image is None:
        raise HTTPException(400, f"Could not decode image: {filename}")
    return image


async def read_image_async(file: UploadFile) -> np.ndarray:
    """Async version of read_image_from_upload."""
    contents = await file.read()
    filename = file.filename or "image.jpg"
    ext = Path(filename).suffix.lower()

    # Handle RAW files
    if ext in RAW_EXTENSIONS:
        if not HAS_RAWPY:
            raise HTTPException(400, f"RAW file support not available. Install rawpy.")
        try:
            with rawpy.imread(io.BytesIO(contents)) as raw:
                rgb = raw.postprocess(
                    use_camera_wb=True,
                    half_size=False,
                    no_auto_bright=False,
                    output_bps=8
                )
                return cv2.cvtColor(rgb, cv2.COLOR_RGB2BGR)
        except Exception as e:
            raise HTTPException(400, f"Could not decode RAW file {filename}: {str(e)}")

    # Standard image formats
    nparr = np.frombuffer(contents, np.uint8)
    image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    if image is None:
        raise HTTPException(400, f"Could not decode image: {filename}")
    return image


def image_to_bytes(image: np.ndarray, format: str = ".jpg", quality: int = 90) -> bytes:
    """Convert OpenCV image to bytes."""
    if format == ".jpg":
        encode_params = [cv2.IMWRITE_JPEG_QUALITY, quality]
    else:
        encode_params = []
    success, encoded = cv2.imencode(format, image, encode_params)
    if not success:
        raise HTTPException(500, "Failed to encode result image")
    return encoded.tobytes()


def merge_brackets_mertens(images: List[np.ndarray]) -> np.ndarray:
    """Merge bracketed exposures using Mertens exposure fusion."""
    if len(images) == 1:
        return images[0]

    # Ensure all images same size
    base_shape = images[0].shape[:2]
    resized = []
    for img in images:
        if img.shape[:2] != base_shape:
            img = cv2.resize(img, (base_shape[1], base_shape[0]))
        resized.append(img)

    # Mertens exposure fusion
    # IMPORTANT: Pass uint8 images directly - Mertens handles conversion internally
    merge_mertens = cv2.createMergeMertens()
    fusion = merge_mertens.process(resized)  # uint8 input, float32 output (0-1 range)
    fusion = np.clip(fusion * 255, 0, 255).astype(np.uint8)

    return fusion


def _build_gaussian_pyramid(img: np.ndarray, levels: int) -> List[np.ndarray]:
    """Build Gaussian pyramid."""
    pyramid = [img.astype(np.float32)]
    for _ in range(levels - 1):
        img = cv2.pyrDown(img)
        pyramid.append(img.astype(np.float32))
    return pyramid


def _build_laplacian_pyramid(img: np.ndarray, levels: int) -> List[np.ndarray]:
    """Build Laplacian pyramid from image."""
    gaussian = _build_gaussian_pyramid(img, levels)
    laplacian = []
    for i in range(levels - 1):
        size = (gaussian[i].shape[1], gaussian[i].shape[0])
        upsampled = cv2.pyrUp(gaussian[i + 1], dstsize=size)
        laplacian.append(gaussian[i] - upsampled)
    laplacian.append(gaussian[-1])  # Top level is just gaussian
    return laplacian


def _reconstruct_from_laplacian(pyramid: List[np.ndarray]) -> np.ndarray:
    """Reconstruct image from Laplacian pyramid."""
    img = pyramid[-1]
    for i in range(len(pyramid) - 2, -1, -1):
        size = (pyramid[i].shape[1], pyramid[i].shape[0])
        img = cv2.pyrUp(img, dstsize=size) + pyramid[i]
    return img


def _compute_weight_map(img: np.ndarray) -> np.ndarray:
    """
    Compute exposure fusion weight map based on:
    - Contrast (Laplacian magnitude)
    - Saturation
    - Well-exposedness (how close to middle gray)
    """
    img_float = img.astype(np.float32) / 255.0

    # Contrast weight (Laplacian filter response)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY).astype(np.float32)
    laplacian = np.abs(cv2.Laplacian(gray, cv2.CV_32F))
    contrast = laplacian / (laplacian.max() + 1e-8)

    # Saturation weight
    saturation = img_float.std(axis=2)

    # Well-exposedness (Gaussian centered at 0.5)
    sigma = 0.2
    well_exposed = np.exp(-0.5 * ((img_float - 0.5) ** 2) / (sigma ** 2))
    well_exposed = np.prod(well_exposed, axis=2)  # Product across channels

    # Combined weight
    weight = (contrast ** 1.0) * (saturation ** 1.0) * (well_exposed ** 1.0)
    weight = weight + 1e-8  # Avoid division by zero

    return weight


def merge_brackets_laplacian(images: List[np.ndarray], levels: int = 6) -> np.ndarray:
    """
    Laplacian pyramid exposure fusion - better than basic Mertens.
    Based on "Exposure Fusion" by Mertens et al. with pyramid blending.
    """
    if len(images) == 1:
        return images[0]

    # Ensure all images same size
    base_shape = images[0].shape[:2]
    resized = []
    for img in images:
        if img.shape[:2] != base_shape:
            img = cv2.resize(img, (base_shape[1], base_shape[0]))
        resized.append(img)

    n_images = len(resized)

    # Compute weight maps for each image
    weights = [_compute_weight_map(img) for img in resized]

    # Normalize weights (sum to 1 at each pixel)
    weight_sum = np.sum(weights, axis=0) + 1e-8
    weights = [w / weight_sum for w in weights]

    # Build Gaussian pyramids of weights
    weight_pyramids = [_build_gaussian_pyramid(w, levels) for w in weights]

    # Build Laplacian pyramids of images
    laplacian_pyramids = [_build_laplacian_pyramid(img, levels) for img in resized]

    # Blend at each pyramid level
    blended_pyramid = []
    for level in range(levels):
        blended_level = np.zeros_like(laplacian_pyramids[0][level])
        for i in range(n_images):
            # Expand weight to 3 channels
            w = weight_pyramids[i][level]
            if len(blended_level.shape) == 3:
                w = np.expand_dims(w, axis=2)
            blended_level += w * laplacian_pyramids[i][level]
        blended_pyramid.append(blended_level)

    # Reconstruct from blended pyramid
    result = _reconstruct_from_laplacian(blended_pyramid)
    result = np.clip(result, 0, 255).astype(np.uint8)

    return result


def merge_brackets_middle_base(images: List[np.ndarray]) -> np.ndarray:
    """
    Professional real estate fusion: Middle exposure as base (neutral balance).

    Based on industry research:
    - Middle exposure provides balanced starting point (no blown highlights, no crushed shadows)
    - Blend BRIGHT image into: dark areas, shadows (for detail)
    - Blend DARK image into: light bulbs and glow areas (for fixture detail)

    No hard-edged spatial masks - only luminosity-based blending for natural results.
    """
    if len(images) == 1:
        return images[0]

    # Ensure all images same size
    base_shape = images[0].shape[:2]
    resized = []
    for img in images:
        if img.shape[:2] != base_shape:
            img = cv2.resize(img, (base_shape[1], base_shape[0]))
        resized.append(img)

    # Sort by average brightness
    brightness = [np.mean(img) for img in resized]
    sorted_indices = np.argsort(brightness)  # Ascending (darkest first)

    darkest = resized[sorted_indices[0]].astype(np.float32)
    brightest = resized[sorted_indices[-1]].astype(np.float32)

    # Middle exposure as base
    if len(resized) >= 3:
        middle_idx = sorted_indices[len(sorted_indices) // 2]
        middle = resized[middle_idx].astype(np.float32)
    else:
        # If only 2 images, blend them 50/50 as "middle"
        middle = (darkest + brightest) / 2

    print(f"   Fusion: Middle base (balanced), bright for shadows, dark for lights")

    # Start with middle exposure
    result = middle.copy()

    # Analyze middle image luminosity
    middle_gray = cv2.cvtColor(middle.astype(np.uint8), cv2.COLOR_BGR2GRAY).astype(np.float32)
    bright_gray = cv2.cvtColor(brightest.astype(np.uint8), cv2.COLOR_BGR2GRAY).astype(np.float32)

    # =========================================================
    # SHADOW/DARK AREA RECOVERY: Blend bright image aggressively
    # =========================================================
    # Where middle exposure is below midtone (<140), pull from bright image
    # This ensures corridor and right side get lifted to match left
    shadow_mask = np.clip((140 - middle_gray) / 100, 0, 1)  # Smooth ramp 40-140
    shadow_mask = cv2.GaussianBlur(shadow_mask, (51, 51), 0)
    shadow_mask = np.stack([shadow_mask] * 3, axis=-1)

    # Strong blend of bright image into darker areas (65%)
    result = result * (1 - shadow_mask * 0.65) + brightest * shadow_mask * 0.65

    # =========================================================
    # GLOBAL BRIGHTNESS: Heavy bright blend for even lighting
    # =========================================================
    # Push bright image hard to get uniform brightness across frame
    global_lift = 0.78  # 78% bright image - bright across entire image
    result = result * (1 - global_lift) + brightest * global_lift

    # =========================================================
    # LIGHT FIXTURE RECOVERY: Blend dark image into blown areas
    # =========================================================
    # Where middle exposure is very bright (>220), use dark image for detail
    highlight_mask = np.clip((middle_gray - 210) / 45, 0, 1)
    highlight_mask = cv2.GaussianBlur(highlight_mask, (31, 31), 0)
    highlight_mask = np.stack([highlight_mask] * 3, axis=-1)

    # Strong blend of dark image into highlights (70% for fixture detail)
    result = result * (1 - highlight_mask * 0.70) + darkest * highlight_mask * 0.70

    # =========================================================
    # EXTREME HIGHLIGHT RECOVERY: Light bulbs specifically
    # =========================================================
    # Where bright image is completely blown (>250), definitely use dark
    extreme_mask = np.clip((bright_gray - 245) / 10, 0, 1)
    extreme_mask = cv2.GaussianBlur(extreme_mask, (21, 21), 0)
    extreme_mask = np.stack([extreme_mask] * 3, axis=-1)

    # Very strong blend for actual light sources (85%)
    result = result * (1 - extreme_mask * 0.85) + darkest * extreme_mask * 0.85

    return np.clip(result, 0, 255).astype(np.uint8)


# ============================================
# ROUTES
# ============================================

@app.get("/")
async def root():
    """API info and capabilities."""
    return {
        "name": "AutoHDR Clone API",
        "version": "1.0.0",
        "status": "running",
        "capabilities": {
            "raw_support": HAS_RAWPY,
            "processor": HAS_PROCESSOR,
            "opencv": CV2_VERSION,
        },
        "supported_formats": {
            "standard": ["jpg", "jpeg", "png", "tiff", "bmp", "webp"],
            "raw": list(RAW_EXTENSIONS) if HAS_RAWPY else []
        },
        "endpoints": {
            "POST /process": "Main processing endpoint",
            "GET /health": "Health check",
            "GET /test": "Test processing pipeline",
            "GET /docs": "API documentation"
        }
    }


@app.get("/health")
async def health():
    """Health check - returns system status."""
    return {
        "status": "healthy",
        "components": {
            "rawpy": {"installed": HAS_RAWPY, "version": RAWPY_VERSION},
            "opencv": {"installed": True, "version": CV2_VERSION},
            "processor": {
                "installed": HAS_PROCESSOR,
                "version": PROCESSOR_VERSION,
                "features": [
                    "CLAHE tone mapping",
                    "Laplacian pyramid fusion",
                    "Cool white balance correction",
                    "Shadow recovery",
                    "S-curve contrast",
                    "Non-local means denoising",
                    "Bilateral edge smoothing",
                    "Local contrast (clarity)",
                    "LAB + HSV color boost",
                ] if HAS_PROCESSOR else []
            },
            "ai_processor": {"installed": HAS_AI_PROCESSOR, "note": "SAM + YOLOv8 + LaMa"},
            "pro_processor": {
                "installed": HAS_PRO_PROCESSOR,
                "version": PRO_VERSION if HAS_PRO_PROCESSOR else None,
                "features": [
                    "Mertens Exposure Fusion",
                    "ECC Bracket Alignment",
                    "Window Detail Recovery",
                    "Flambient Tone Mapping",
                    "Professional Window Pull",
                    "Edge-Aware Processing",
                ] if HAS_PRO_PROCESSOR else []
            },
        },
        "raw_formats_supported": len(RAW_EXTENSIONS) if HAS_RAWPY else 0,
        "pro_processor_available": HAS_PRO_PROCESSOR,
        "quality_level": "95%+ AutoHDR" if HAS_PRO_PROCESSOR else ("90% AutoHDR" if HAS_AI_PROCESSOR else "60% AutoHDR"),
    }


@app.get("/test")
async def test_processing():
    """
    Test the processing pipeline with a generated gradient image.
    Returns a processed test image to verify the pipeline works.
    """
    if not HAS_PROCESSOR:
        raise HTTPException(500, "Processor not available")

    try:
        start_time = time.time()

        # Create a test gradient image (simulates a photo with shadows and highlights)
        height, width = 480, 640
        test_image = np.zeros((height, width, 3), dtype=np.uint8)

        # Gradient background (sky simulation)
        for y in range(height):
            ratio = y / height
            test_image[y, :] = [
                int(255 * (1 - ratio * 0.5)),  # Blue (sky to ground)
                int(200 * (1 - ratio * 0.3)),  # Green
                int(180 * (1 - ratio * 0.4)),  # Red
            ]

        # Add some "windows" (bright rectangles)
        cv2.rectangle(test_image, (100, 200), (200, 350), (255, 255, 255), -1)
        cv2.rectangle(test_image, (450, 180), (550, 320), (255, 255, 255), -1)

        # Add "shadows" (dark areas)
        cv2.rectangle(test_image, (250, 350), (400, 450), (30, 30, 30), -1)

        # Process with default settings
        processor = AutoHDRProcessor(ProcessingSettings())
        result = processor.process(test_image)

        elapsed_ms = (time.time() - start_time) * 1000

        # Encode as JPEG
        _, buffer = cv2.imencode('.jpg', result, [cv2.IMWRITE_JPEG_QUALITY, 85])

        return Response(
            content=buffer.tobytes(),
            media_type="image/jpeg",
            headers={
                "Content-Disposition": "inline; filename=test_result.jpg",
                "X-Processing-Time-Ms": str(round(elapsed_ms, 2)),
            }
        )

    except Exception as e:
        raise HTTPException(500, f"Test failed: {str(e)}")


# ============================================
# MAIN PROCESSING ENDPOINT (matches Vercel frontend)
# ============================================

@app.post("/process")
async def process_images(
    images: List[UploadFile] = File(..., description="Images to process"),
    mode: str = Query("hdr", description="Processing mode: hdr or twilight"),
    brightness: float = Query(0, ge=-2, le=2),
    contrast: float = Query(0, ge=-2, le=2),
    vibrance: float = Query(0, ge=-2, le=2),
    whiteBalance: float = Query(0, ge=-2, le=2),
    ai: bool = Query(False, description="Use AI-enhanced processing (SAM, YOLOv8, LaMa)"),
    grass: bool = Query(False, description="Enhance grass (vibrant green)"),
    signs: bool = Query(False, description="Remove signs (requires ai=true for best results)"),
):
    """
    Main processing endpoint - matches Vercel frontend API.

    - **HDR mode**: Upload 2-9 bracketed exposures, merges with Mertens fusion
    - **Twilight mode**: Upload 1 daytime photo, converts to dusk
    - **AI mode**: Use SAM + YOLOv8 + LaMa for 90% AutoHDR quality

    Supports all RAW formats: ARW, CR2, NEF, DNG, etc.
    """
    start_time = time.time()

    # Validate
    if not images:
        raise HTTPException(400, "No images provided")

    if not HAS_PROCESSOR:
        raise HTTPException(500, "Processor not available - check server logs")

    # Log request
    filenames = [img.filename for img in images]
    print(f"📸 Processing {len(images)} images: {filenames}")
    print(f"   Mode: {mode}, Settings: b={brightness}, c={contrast}, v={vibrance}, wb={whiteBalance}")

    try:
        # ==========================================
        # STEP 1: Read all images (including RAW)
        # ==========================================
        image_arrays = []
        for i, upload in enumerate(images):
            print(f"   Reading [{i+1}/{len(images)}]: {upload.filename}")
            try:
                img = await read_image_async(upload)
                print(f"   ✓ Loaded: {img.shape[1]}x{img.shape[0]} pixels")
                image_arrays.append(img)
            except HTTPException:
                raise
            except Exception as e:
                raise HTTPException(400, f"Failed to read {upload.filename}: {str(e)}")

        # ==========================================
        # STEP 2: Process with Pro Processor (if available and multiple brackets)
        # ==========================================
        if len(image_arrays) > 1 and HAS_PRO_PROCESSOR:
            pro_settings = ProSettings(
                brightness=brightness,
                contrast=contrast,
                vibrance=vibrance,
                white_balance=whiteBalance,
            )
            pro_processor = AutoHDRProProcessor(pro_settings)

            # Check if this looks like multiple scenes (more than typical bracket count)
            if len(image_arrays) > 5:
                # Multi-scene processing: auto-group by scene
                print(f"   🎯 Multi-scene mode: {len(image_arrays)} images, auto-grouping...")
                results = pro_processor.process_multiple_scenes(image_arrays, auto_group=True)

                if len(results) == 1:
                    # Single scene detected
                    result = results[0]
                    result_bytes = image_to_bytes(result, ".jpg", quality=98)
                    elapsed_ms = (time.time() - start_time) * 1000
                    print(f"   ✓ Single scene processed in {elapsed_ms:.0f}ms")

                    return Response(
                        content=result_bytes,
                        media_type="image/jpeg",
                        headers={
                            "Content-Disposition": f'attachment; filename="hdrit_{mode}_1.jpg"',
                            "X-Processing-Time-Ms": str(round(elapsed_ms, 2)),
                            "X-Images-Processed": str(len(images)),
                            "X-Scenes-Detected": "1",
                            "X-Processor": f"Pro v{PRO_VERSION}",
                        }
                    )
                else:
                    # Multiple scenes - return as ZIP
                    import zipfile
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                        for i, result in enumerate(results):
                            img_bytes = image_to_bytes(result, ".jpg", quality=98)
                            zf.writestr(f"hdrit_scene_{i+1:02d}.jpg", img_bytes)

                    zip_buffer.seek(0)
                    elapsed_ms = (time.time() - start_time) * 1000
                    print(f"   ✓ {len(results)} scenes processed in {elapsed_ms:.0f}ms")

                    return Response(
                        content=zip_buffer.getvalue(),
                        media_type="application/zip",
                        headers={
                            "Content-Disposition": f'attachment; filename="hdrit_batch_{len(results)}_scenes.zip"',
                            "X-Processing-Time-Ms": str(round(elapsed_ms, 2)),
                            "X-Images-Processed": str(len(images)),
                            "X-Scenes-Detected": str(len(results)),
                            "X-Processor": f"Pro v{PRO_VERSION}",
                        }
                    )
            else:
                # Standard bracket processing (2-5 images = single scene)
                print(f"   🔥 Using Pro Processor v{PRO_VERSION} for {len(image_arrays)} brackets...")
                result = pro_processor.process_brackets(image_arrays)

                # Encode and return directly - Pro processor handles everything
                result_bytes = image_to_bytes(result, ".jpg", quality=95)
                elapsed_ms = (time.time() - start_time) * 1000
                print(f"   ✓ Pro Processor complete in {elapsed_ms:.0f}ms")
                print(f"   📦 Response size: {len(result_bytes) / 1024:.1f} KB")

                return Response(
                    content=result_bytes,
                    media_type="image/jpeg",
                    headers={
                        "Content-Disposition": f'attachment; filename="hdrit_{mode}.jpg"',
                        "Content-Length": str(len(result_bytes)),
                        "X-Processing-Time-Ms": str(round(elapsed_ms, 2)),
                        "X-Images-Processed": str(len(images)),
                        "X-Processor": f"Pro v{PRO_VERSION}",
                        "Access-Control-Allow-Origin": "*",
                        "Access-Control-Expose-Headers": "Content-Length, X-Processing-Time-Ms, X-Processor",
                        "Cache-Control": "no-cache",
                    }
                )

        # Fallback: merge brackets with legacy fusion
        if len(image_arrays) > 1:
            print(f"   Merging {len(image_arrays)} brackets with middle-base fusion...")
            base_image = merge_brackets_middle_base(image_arrays)
        else:
            base_image = image_arrays[0]

        # ==========================================
        # STEP 3: Process (Basic or AI-Enhanced)
        # ==========================================
        if ai and HAS_AI_PROCESSOR:
            print("   🤖 Using AI-enhanced processing (SAM + YOLOv8 + LaMa)...")
            ai_settings = AIProcessingSettings(
                brightness=brightness,
                contrast=contrast,
                vibrance=vibrance,
                white_balance=whiteBalance,
                twilight_style="pink" if mode == "twilight" else None,
                grass_enhancement=grass,
                sign_removal=signs,
                use_ai_segmentation=True,
                use_ai_detection=True,
                use_ai_inpainting=signs,  # Only use LaMa if removing signs
            )
            processor = AIEnhancedProcessor(ai_settings)
            result = processor.process(base_image)
        else:
            if ai and not HAS_AI_PROCESSOR:
                print("   ⚠️ AI processor not available, using basic processing")
            print("   Applying HDR processing...")
            settings = ProcessingSettings(
                brightness=brightness,
                contrast=contrast,
                vibrance=vibrance,
                white_balance=whiteBalance,
                twilight_style="pink" if mode == "twilight" else None,
                grass_replacement=grass,
                sign_removal=signs,
            )
            processor = AutoHDRProcessor(settings)
            result = processor.process(base_image)

        # ==========================================
        # STEP 3: Encode and return
        # ==========================================
        result_bytes = image_to_bytes(result, ".jpg", quality=98)  # Full resolution output

        elapsed_ms = (time.time() - start_time) * 1000
        print(f"   ✓ Complete in {elapsed_ms:.0f}ms, output: {len(result_bytes)} bytes")

        return Response(
            content=result_bytes,
            media_type="image/jpeg",
            headers={
                "Content-Disposition": f'attachment; filename="autohdr_{mode}.jpg"',
                "X-Processing-Time-Ms": str(round(elapsed_ms, 2)),
                "X-Images-Processed": str(len(images)),
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"   ✗ ERROR: {str(e)}")
        traceback.print_exc()
        raise HTTPException(500, f"Processing failed: {str(e)}")


# ============================================
# LEGACY ENDPOINTS (for direct API use)
# ============================================

@app.post("/hdr/merge")
async def hdr_merge(
    images: List[UploadFile] = File(...),
    method: str = Form("mertens"),
    tone_map: str = Form("reinhard"),
    align: bool = Form(True)
):
    """
    Merge multiple exposure brackets into HDR.

    Upload 2-9 images taken at different exposures.
    Returns the merged, tone-mapped result.
    """
    if len(images) < 2:
        raise HTTPException(400, "Need at least 2 images for HDR merge")
    if len(images) > 9:
        raise HTTPException(400, "Maximum 9 images supported")

    # Read all images (with RAW support)
    loaded_images = []
    for img_file in images:
        img = await read_image_async(img_file)
        loaded_images.append(img)

    # Use Mertens fusion (built-in, always works)
    result = merge_brackets_mertens(loaded_images)

    # Apply HDR processing
    processor = AutoHDRProcessor(ProcessingSettings())
    result = processor.process(result)

    # Return as JPEG
    result_bytes = image_to_bytes(result, ".jpg", quality=90)

    return StreamingResponse(
        io.BytesIO(result_bytes),
        media_type="image/jpeg",
        headers={"Content-Disposition": "attachment; filename=hdr_result.jpg"}
    )


@app.post("/effects/day-to-dusk")
async def day_to_dusk(
    image: UploadFile = File(...),
    style: str = Form("pink"),  # pink, blue, orange
    brightness: float = Form(-0.5),  # -2 to 2
):
    """
    Convert daytime exterior photo to twilight/dusk.

    Parameters:
    - style: Twilight color (pink, blue, orange)
    - brightness: Adjustment (-2 to 2, negative = darker)
    """
    # Read image (with RAW support)
    img = await read_image_async(image)

    # Configure for twilight
    settings = ProcessingSettings(
        brightness=brightness,
        twilight_style=style if style in ["pink", "blue", "orange"] else "pink"
    )

    # Process
    processor = AutoHDRProcessor(settings)
    result = processor.process(img)

    # Return as JPEG
    result_bytes = image_to_bytes(result, ".jpg")

    return StreamingResponse(
        io.BytesIO(result_bytes),
        media_type="image/jpeg",
        headers={"Content-Disposition": "attachment; filename=twilight_result.jpg"}
    )


@app.post("/edit/remove")
async def remove_object(
    image: UploadFile = File(...),
    prompt: str = Form(None),
    mask: UploadFile = File(None)
):
    """
    Remove objects from image using AI inpainting.

    Provide either:
    - prompt: Text description of what to remove (e.g., "remove the car")
    - mask: Binary mask image where white = area to remove

    Coming soon: Requires LaMa or IOPaint integration.
    """
    # TODO: Implement with LaMa / Grounded-SAM
    raise HTTPException(501, "Object removal coming soon. Use IOPaint for now: https://github.com/Sanster/IOPaint")


@app.post("/edit/sky-replace")
async def sky_replace(
    image: UploadFile = File(...),
    sky_type: str = Form("twilight")  # twilight, blue, dramatic, cloudy
):
    """
    Replace sky in image.

    Coming soon: Requires sky segmentation model.
    """
    # TODO: Implement with sky segmentation
    raise HTTPException(501, "Sky replacement coming soon. Requires segmentation model.")


# ============================================
# OBSERVER AUTH PROXY
# ============================================
# Proxy Observer Auth requests to port 8001
# so everything works through a single ngrok tunnel

import httpx

OBSERVER_URL = "http://localhost:8001"

@app.post("/observer/start")
async def observer_start(request: Request):
    """Proxy auth start to Observer."""
    try:
        body = await request.body()
        async with httpx.AsyncClient() as client:
            resp = await client.post(f"{OBSERVER_URL}/auth/start", content=body,
                                     headers={"Content-Type": "application/json"}, timeout=10)
            return JSONResponse(resp.json(), status_code=resp.status_code)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=502)

@app.get("/observer/check/{session_id}")
async def observer_check(session_id: str):
    """Proxy auth check to Observer."""
    try:
        async with httpx.AsyncClient() as client:
            resp = await client.get(f"{OBSERVER_URL}/auth/check/{session_id}", timeout=10)
            return JSONResponse(resp.json(), status_code=resp.status_code)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=502)

@app.post("/observer/approve/{session_id}")
async def observer_approve(session_id: str, request: Request):
    """Proxy auth approve to Observer."""
    try:
        body = await request.body()
        async with httpx.AsyncClient() as client:
            resp = await client.post(f"{OBSERVER_URL}/auth/approve/{session_id}", content=body,
                                     headers={"Content-Type": "application/json"}, timeout=10)
            return JSONResponse(resp.json(), status_code=resp.status_code)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=502)

@app.post("/observer/deny/{session_id}")
async def observer_deny(session_id: str, request: Request):
    """Proxy auth deny to Observer."""
    try:
        body = await request.body()
        async with httpx.AsyncClient() as client:
            resp = await client.post(f"{OBSERVER_URL}/auth/deny/{session_id}", content=body,
                                     headers={"Content-Type": "application/json"}, timeout=10)
            return JSONResponse(resp.json(), status_code=resp.status_code)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=502)

@app.post("/observer/code")
async def observer_code(request: Request):
    """Proxy code-based approval to Observer."""
    try:
        body = await request.body()
        async with httpx.AsyncClient() as client:
            resp = await client.post(f"{OBSERVER_URL}/auth/code", content=body,
                                     headers={"Content-Type": "application/json"}, timeout=10)
            return JSONResponse(resp.json(), status_code=resp.status_code)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=502)

@app.get("/observer/scan/{session_id}")
async def observer_scan(session_id: str, request: Request):
    """QR scan landing - auto-approves from phone with device_id."""
    device_id = request.query_params.get("device_id", "")
    device_name = request.query_params.get("device_name", "Phone Scanner")
    if device_id:
        try:
            async with httpx.AsyncClient() as client:
                resp = await client.post(f"{OBSERVER_URL}/auth/approve/{session_id}",
                    json={"device_id": device_id, "device_name": device_name}, timeout=10)
                data = resp.json()
                if data.get("success"):
                    return JSONResponse({"status": "approved", "message": "Access granted!"})
        except:
            pass
    # Return a simple page that sends device info and approves
    from fastapi.responses import HTMLResponse
    return HTMLResponse(f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Observer - Approve Login</title>
<style>body{{background:#000;color:#0f0;font-family:system-ui;display:flex;align-items:center;justify-content:center;height:100vh;margin:0}}
.card{{text-align:center;padding:30px;border:1px solid #0f03;border-radius:16px;background:#0f01}}
button{{background:#0f0;color:#000;border:none;padding:15px 40px;border-radius:12px;font-size:18px;font-weight:bold;cursor:pointer;margin-top:20px}}
.deny{{background:#f00;color:#fff;margin-left:10px}}</style></head>
<body><div class="card">
<h1>🔐 Login Request</h1>
<p>Approve access to AXE?</p>
<button onclick="approve()">✓ Approve</button>
<button class="deny" onclick="deny()">✗ Deny</button>
<p id="status" style="margin-top:20px"></p>
<script>
const sid="{session_id}";
let did=localStorage.getItem('observer_device_id');
if(!did){{did=crypto.randomUUID();localStorage.setItem('observer_device_id',did)}}
async function approve(){{
  const r=await fetch('/observer/approve/'+sid,{{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify({{device_id:did,device_name:navigator.userAgent.slice(0,30)}})}});
  const d=await r.json();
  document.getElementById('status').textContent=d.success?'✓ Approved!':'Error: '+(d.error||'failed');
}}
async function deny(){{
  await fetch('/observer/deny/'+sid,{{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify({{device_id:did}})}});
  document.getElementById('status').textContent='✗ Denied';
}}
</script></div></body></html>""")

@app.get("/observer/app")
async def observer_app_redirect():
    """Redirect to Observer PWA."""
    from fastapi.responses import RedirectResponse
    return RedirectResponse("http://localhost:8001/app")


# ============================================
# OLLAMA PROXY (for Klaus frontend)
# ============================================

import httpx
import json as _json
import uuid
import re as _re

OLLAMA_BASE = "http://localhost:11434"
_ollama_client = httpx.AsyncClient(timeout=httpx.Timeout(300.0, connect=10.0))

@app.api_route("/ollama/{path:path}", methods=["GET", "POST", "PUT", "DELETE"])
async def ollama_proxy(path: str, request: Request):
    """Proxy requests to local Ollama instance."""
    url = f"{OLLAMA_BASE}/{path}"
    body = await request.body()

    is_stream = False
    if body:
        try:
            is_stream = _json.loads(body).get("stream", False)
        except Exception:
            pass

    if is_stream:
        req = _ollama_client.build_request(request.method, url, content=body,
                                            headers={"content-type": "application/json"})
        resp = await _ollama_client.send(req, stream=True)
        async def stream_response():
            try:
                async for line in resp.aiter_lines():
                    yield line.encode() + b"\n"
            finally:
                await resp.aclose()
        return StreamingResponse(stream_response(), status_code=resp.status_code,
                                 media_type="application/x-ndjson")
    else:
        resp = await _ollama_client.request(request.method, url, content=body,
                                             headers={"content-type": "application/json"})
        return Response(content=resp.content, status_code=resp.status_code,
                       media_type=resp.headers.get("content-type", "application/json"))

def _load_team_context(max_messages: int = 30) -> str:
    """Load recent team channel messages for Klaus memory context."""
    channel_path = os.path.expanduser("~/Desktop/M1transfer/axe-memory/team/channel.jsonl")
    context_path = os.path.expanduser("~/.axe/memory/active_context.json")
    lines = []
    try:
        with open(channel_path, "r") as f:
            all_lines = f.readlines()
            for raw in all_lines[-max_messages:]:
                try:
                    m = _json.loads(raw.strip())
                    lines.append(f"[{m.get('from','?')}→{m.get('to','team')}] {m.get('msg','')[:200]}")
                except Exception:
                    pass
    except Exception:
        pass
    context_parts = []
    if lines:
        context_parts.append("Recent team channel (Forge, Cortana, Klaus, James):\n" + "\n".join(lines))
    try:
        with open(context_path, "r") as f:
            ctx = _json.load(f)
            if ctx.get("current_projects"):
                context_parts.append("Active projects: " + _json.dumps(ctx["current_projects"][:5]))
            if ctx.get("critical_facts"):
                context_parts.append("Critical facts: " + _json.dumps(ctx["critical_facts"][:5]))
    except Exception:
        pass
    return "\n\n".join(context_parts)

# === Team Channel API ===
TEAM_CHANNEL_PATH = os.path.expanduser("~/Desktop/M1transfer/axe-memory/team/channel.jsonl")

class TeamMessage(BaseModel):
    from_agent: str = "unknown"
    to: str = "team"
    msg: str
    type: str = "message"

@app.post("/team/message")
async def post_team_message(payload: TeamMessage):
    """Post a message to the team channel. Used by all agents for comms."""
    import datetime as _dt
    entry = {
        "ts": _dt.datetime.now(_dt.timezone.utc).isoformat(),
        "from": payload.from_agent,
        "to": payload.to,
        "type": payload.type,
        "msg": payload.msg,
    }
    with open(TEAM_CHANNEL_PATH, "a") as f:
        f.write(_json.dumps(entry) + "\n")
    return {"status": "sent", "ts": entry["ts"]}

@app.get("/team/messages")
async def get_team_messages(limit: int = 20):
    """Read recent team channel messages."""
    try:
        with open(TEAM_CHANNEL_PATH, "r") as f:
            all_lines = f.readlines()
        messages = []
        for raw in all_lines[-limit:]:
            try:
                messages.append(_json.loads(raw.strip()))
            except:
                pass
        return {"messages": messages}
    except FileNotFoundError:
        return {"messages": []}


_IMI_MEMORY_DIR = os.path.expanduser("~/.axe/memory/imi_conversations")
os.makedirs(_IMI_MEMORY_DIR, exist_ok=True)

@app.post("/klaus/imi/save")
async def klaus_imi_save(request: Request):
    """Auto-save IMI conversation to memory."""
    body = await request.json()
    conv_id = body.get("conversation_id", "unknown")
    title = body.get("title", "untitled")
    messages = body.get("messages", [])
    filepath = os.path.join(_IMI_MEMORY_DIR, f"{conv_id}.json")
    with open(filepath, "w") as f:
        _json.dump({"id": conv_id, "title": title, "messages": messages, "saved_at": datetime.now().isoformat()}, f, indent=2)
    return {"status": "saved", "path": filepath}

def _parse_crosstab_xlsx(file_bytes: bytes) -> dict:
    """Parse IMI crosstab xlsx (segments as columns, Q-blocks as rows)."""
    import openpyxl, io as _io
    wb = openpyxl.load_workbook(_io.BytesIO(file_bytes))
    ws = wb.active
    seg_headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=3, column=col).value
        if val and str(val).strip():
            seg_headers[col] = str(val).strip()
    questions = []
    current_q = None
    for row in range(4, ws.max_row + 1):
        first = ws.cell(row=row, column=1).value
        first_str = str(first).strip() if first else ''
        if first_str.startswith('Q') and any(c.isdigit() for c in first_str[:3]) and ' ' in first_str:
            if current_q:
                questions.append(current_q)
            current_q = {'id': first_str.split(' ')[0], 'text': first_str, 'options': [], 'base_sizes': {}}
            continue
        if 'sample' in first_str.lower():
            if current_q:
                for col, seg in seg_headers.items():
                    v = ws.cell(row=row, column=col).value
                    if v:
                        try: current_q['base_sizes'][seg] = int(float(v))
                        except: pass
            continue
        if first_str and current_q:
            vals = {}
            has_num = False
            for col, seg in seg_headers.items():
                v = ws.cell(row=row, column=col).value
                if v is not None:
                    try:
                        vals[seg] = float(v)
                        has_num = True
                    except: pass
            if has_num:
                current_q['options'].append({'label': first_str, 'values': vals})
    if current_q:
        questions.append(current_q)
    total_n = 0
    for q in questions:
        if q.get('base_sizes') and 'Canada' in q['base_sizes']:
            total_n = q['base_sizes']['Canada']
            break
    return {'total_n': total_n, 'segments': list(seg_headers.values()), 'questions': questions, 'format': 'crosstab'}


def _parse_flat_xlsx(file_bytes: bytes) -> dict:
    """Parse flat xlsx table (rows=items, cols=metrics). E.g. brand health tracker."""
    import openpyxl, io as _io
    wb = openpyxl.load_workbook(_io.BytesIO(file_bytes))
    ws = wb.active
    headers = [str(ws.cell(row=1, column=c).value or '').strip() for c in range(1, ws.max_column + 1)]
    label_col = 0  # first column is the item label
    metric_cols = []
    sample_col = None
    for i, h in enumerate(headers):
        if i == 0:
            continue
        hl = h.lower()
        if 'sample' in hl or 'n' == hl:
            sample_col = i
        elif h:
            metric_cols.append(i)
    # Build one "question" per metric column, with each row as an option
    questions = []
    rows_data = []
    for row in range(2, ws.max_row + 1):
        label = ws.cell(row=row, column=1).value
        if not label:
            continue
        row_vals = {}
        for ci in metric_cols:
            v = ws.cell(row=row, column=ci + 1).value
            if v is not None:
                try: row_vals[headers[ci]] = float(v)
                except: row_vals[headers[ci]] = v
        n = None
        if sample_col is not None:
            sv = ws.cell(row=row, column=sample_col + 1).value
            if sv:
                try: n = int(float(sv))
                except: pass
        rows_data.append({'label': str(label).strip(), 'values': row_vals, 'n': n})
    # Create one question per numeric metric
    for ci in metric_cols:
        metric_name = headers[ci]
        opts = []
        base_sizes = {}
        for rd in rows_data:
            v = rd['values'].get(metric_name)
            if v is not None and isinstance(v, (int, float)):
                opts.append({'label': rd['label'], 'values': {'Total': v}})
                if rd.get('n'):
                    base_sizes[rd['label']] = rd['n']
        if opts:
            questions.append({
                'id': metric_name,
                'text': metric_name,
                'options': opts,
                'base_sizes': base_sizes
            })
    total_n = max((rd.get('n') or 0 for rd in rows_data), default=0)
    return {'total_n': total_n, 'segments': ['Total'], 'questions': questions, 'format': 'flat', 'items': [rd['label'] for rd in rows_data]}


def _parse_flat_csv(file_bytes: bytes) -> dict:
    """Parse flat CSV (rows=respondents or items, cols=metrics)."""
    import csv, io as _io
    text = file_bytes.decode('utf-8', errors='replace')
    reader = csv.DictReader(_io.StringIO(text))
    rows = list(reader)
    if not rows:
        return {'total_n': 0, 'segments': [], 'questions': [], 'format': 'csv_empty'}
    headers = list(rows[0].keys())
    # Detect if respondent-level (has ID/respondent col) or item-level (brand per row)
    first_h = headers[0].lower()
    is_respondent = any(k in first_h for k in ['respondent', 'id', 'resp'])
    if is_respondent:
        # Aggregate respondent-level data: compute means/distributions per segment
        # Find segment columns (categorical) vs metric columns (numeric)
        seg_cols = []
        metric_cols = []
        for h in headers:
            # Try first 20 rows to detect type
            numeric_count = 0
            for r in rows[:20]:
                try:
                    float(r[h])
                    numeric_count += 1
                except:
                    pass
            if numeric_count > 10:
                metric_cols.append(h)
            elif h.lower() not in ['respondent_id', 'id', 'survey_date', 'survey_method']:
                # Check if it's a useful segment (few unique values)
                uniques = set(r[h] for r in rows[:100] if r.get(h))
                if 2 <= len(uniques) <= 20:
                    seg_cols.append(h)
        # Build questions from metric columns, segmented by seg_cols
        questions = []
        for mc in metric_cols:
            opts = []
            # Compute overall mean
            vals = []
            for r in rows:
                try: vals.append(float(r[mc]))
                except: pass
            if not vals:
                continue
            overall_mean = sum(vals) / len(vals)
            # For each segment column, compute means per group
            seg_data = {}
            for sc in seg_cols[:5]:  # limit to top 5 segments
                groups = {}
                for r in rows:
                    grp = r.get(sc, '')
                    if not grp:
                        continue
                    try:
                        v = float(r[mc])
                        groups.setdefault(grp, []).append(v)
                    except:
                        pass
                for grp, gvals in groups.items():
                    seg_data[f"{sc}: {grp}"] = round(sum(gvals) / len(gvals), 3)
            seg_data['Total'] = round(overall_mean, 3)
            questions.append({
                'id': mc,
                'text': mc.replace('_', ' ').title(),
                'options': [{'label': 'Mean Score', 'values': seg_data}],
                'base_sizes': {'Total': len(vals)}
            })
        return {'total_n': len(rows), 'segments': ['Total'] + seg_cols[:5], 'questions': questions, 'format': 'respondent_csv'}
    else:
        # Item-level CSV (like brand per row) — same logic as flat xlsx
        metric_cols = []
        sample_col = None
        for h in headers[1:]:
            hl = h.lower()
            if 'sample' in hl or h == 'N' or h == 'n':
                sample_col = h
            else:
                # Check if numeric
                try:
                    float(rows[0].get(h, ''))
                    metric_cols.append(h)
                except:
                    pass
        rows_data = []
        for r in rows:
            label = r[headers[0]]
            vals = {}
            for mc in metric_cols:
                try: vals[mc] = float(r[mc])
                except: vals[mc] = r.get(mc)
            n = None
            if sample_col:
                try: n = int(float(r[sample_col]))
                except: pass
            rows_data.append({'label': label, 'values': vals, 'n': n})
        questions = []
        for mc in metric_cols:
            opts = []
            base_sizes = {}
            for rd in rows_data:
                v = rd['values'].get(mc)
                if v is not None and isinstance(v, (int, float)):
                    opts.append({'label': rd['label'], 'values': {'Total': v}})
                    if rd.get('n'):
                        base_sizes[rd['label']] = rd['n']
            if opts:
                questions.append({'id': mc, 'text': mc.replace('_', ' ').title(), 'options': opts, 'base_sizes': base_sizes})
        total_n = max((rd.get('n') or 0 for rd in rows_data), default=0)
        return {'total_n': total_n, 'segments': ['Total'], 'questions': questions, 'format': 'item_csv', 'items': [rd['label'] for rd in rows_data]}


def _parse_json_survey(file_bytes: bytes) -> dict:
    """Parse JSON survey data (nested structure)."""
    data = _json.loads(file_bytes.decode('utf-8', errors='replace'))
    # If it's already in our format, use directly
    if 'questions' in data and 'total_n' in data:
        data['format'] = 'native_json'
        return data
    # Otherwise it's a nested dict (e.g. campaigns with metrics)
    questions = []
    items = list(data.keys()) if isinstance(data, dict) else []
    if isinstance(data, dict):
        # Flatten nested dict into questions
        # Each top-level key is an item, nested values are metrics
        def _flatten(prefix, obj, result):
            if isinstance(obj, dict):
                for k, v in obj.items():
                    _flatten(f"{prefix}.{k}" if prefix else k, v, result)
            elif isinstance(obj, (int, float)):
                result[prefix] = obj
        all_metrics = {}
        for item_key, item_data in data.items():
            flat = {}
            _flatten('', item_data, flat)
            all_metrics[item_key] = flat
        # Get union of all metric keys
        all_keys = set()
        for m in all_metrics.values():
            all_keys.update(m.keys())
        for mk in sorted(all_keys):
            opts = []
            for item_key in items:
                v = all_metrics.get(item_key, {}).get(mk)
                if v is not None:
                    opts.append({'label': item_key, 'values': {'Total': v}})
            if opts:
                questions.append({'id': mk, 'text': mk.replace('.', ' > ').replace('_', ' ').title(), 'options': opts, 'base_sizes': {}})
    return {'total_n': 0, 'segments': ['Total'], 'questions': questions, 'format': 'nested_json', 'items': items}


@app.post("/klaus/imi/upload-survey")
async def imi_upload_survey(file: UploadFile):
    """Upload survey file (xlsx/csv/json), parse it, store it, return structured data."""
    import io as _io
    fname = file.filename or ''
    if not fname.endswith(('.xlsx', '.xls', '.csv', '.json')):
        raise HTTPException(status_code=400, detail="Must be .xlsx, .csv, or .json")
    file_bytes = await file.read()
    survey_name = fname.rsplit('.', 1)[0].replace('_', ' ').strip()
    if 'book' in survey_name.lower():
        survey_name = 'Canadian Sports Fandom Pulse'

    try:
        if fname.endswith(('.xlsx', '.xls')):
            # Try crosstab format first (has Q-blocks starting row 4)
            parsed = _parse_crosstab_xlsx(file_bytes)
            if not parsed['questions']:
                # Fallback to flat table format
                parsed = _parse_flat_xlsx(file_bytes)
        elif fname.endswith('.csv'):
            parsed = _parse_flat_csv(file_bytes)
        elif fname.endswith('.json'):
            parsed = _parse_json_survey(file_bytes)
        else:
            raise HTTPException(status_code=400, detail="Unsupported format")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Parse error: {str(e)}")

    survey_data = {
        'survey_name': survey_name,
        'total_n': parsed.get('total_n', 0),
        'segments': parsed.get('segments', []),
        'questions': parsed.get('questions', []),
        'format': parsed.get('format', 'unknown'),
        'items': parsed.get('items', [])
    }
    sid = survey_name.lower().replace(' ', '_')
    _SURVEY_STORE[sid] = {'structured_data': survey_data, 'raw_text': _json.dumps(survey_data, default=str)}
    return {"survey_id": sid, "survey_name": survey_name, "total_n": survey_data['total_n'],
            "questions_found": len(survey_data['questions']), "segments": survey_data['segments'],
            "format": survey_data['format']}


@app.post("/klaus/imi/generate-deck")
async def imi_generate_deck(request: Request):
    """Generate an IMI-branded PPTX deck from a loaded survey."""
    body = await request.json()
    survey_id = body.get("survey_id")
    if not survey_id or survey_id not in _SURVEY_STORE:
        raise HTTPException(status_code=404, detail=f"Survey '{survey_id}' not loaded. Available: {list(_SURVEY_STORE.keys())}")
    survey = _SURVEY_STORE[survey_id]
    survey_data = survey.get("structured_data")
    if not survey_data:
        raise HTTPException(status_code=400, detail="No structured data for this survey")
    from imi_deck_generator import generate_imi_deck
    pptx_bytes = generate_imi_deck(survey_data)
    fname = survey_data.get("survey_name", "IMI_Analysis").replace(" ", "_") + "_IMI.pptx"
    return Response(
        content=pptx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )


@app.get("/klaus/imi/surveys")
async def imi_list_surveys():
    """List loaded surveys."""
    return {"surveys": [
        {"id": sid, "name": s.get("structured_data", {}).get("survey_name", sid),
         "total_n": s.get("structured_data", {}).get("total_n", 0)}
        for sid, s in _SURVEY_STORE.items()
    ]}


@app.post("/klaus/imi/chat")
async def klaus_imi_chat(request: Request):
    """Klaus IMI chat endpoint - proxies to Ollama with klaus-imi model, with team memory."""
    body = await request.json()
    message = body.get("message", "")
    history = body.get("history", [])

    # Check for pre-baked demo responses (zero API spend)
    try:
        from imi_cached_responses import CACHED_RESPONSES
        if message in CACHED_RESPONSES and not history:
            cached = CACHED_RESPONSES[message]
            import asyncio
            async def stream_cached():
                # Simulate Ollama streaming format for frontend compatibility
                words = cached.split(' ')
                chunk_size = 3
                for i in range(0, len(words), chunk_size):
                    chunk = ' '.join(words[i:i+chunk_size])
                    if i > 0:
                        chunk = ' ' + chunk
                    yield _json.dumps({"model": "klaus-imi", "message": {"role": "assistant", "content": chunk}, "done": False}).encode() + b"\n"
                    await asyncio.sleep(0.02)
                yield _json.dumps({"model": "klaus-imi", "message": {"role": "assistant", "content": ""}, "done": True}).encode() + b"\n"
            return StreamingResponse(stream_cached(), status_code=200, media_type="application/x-ndjson")
    except ImportError:
        pass  # No cached responses file, continue to Ollama
    survey_id = body.get("survey_id")

    # Inject team memory context as system message
    team_context = _load_team_context()
    system_msg = """You are Klaus, the AI intelligence engine for IMI International (consultimi.com) — a global consumer research firm with 55+ years of data across 18 countries covering 70% of global GDP. You serve as a senior insight consultant embedded in IMI's proprietary platform.

CARDINAL RULES:
- NEVER fabricate data. Only cite numbers from the provided survey data context.
- If data isn't loaded, say so. Never guess or use "typical" numbers.
- Lead with the insight, not the methodology. You speak like a senior consultant presenting to a CMO.
- Quantify everything: "3.4x more likely" not "much more likely"
- Every insight must answer "so what?" — connect data to business decisions.
- Use bold for key numbers. Keep responses focused and actionable.
- Reference "IMI Pulse™ data" and "Say-Do Gap™" when relevant.
- Always end strategic analyses with a SO WHAT? section.
- When presenting rankings use: "**27%** — FIFA World Cup" format.
- For derived metrics (gaps, index scores), show your math.

RESPONSE STRUCTURE FOR ANALYSIS:
1. The Landscape — top-line national results
2. The Key Divide — most dramatic segment split
3. The Generational Lens — age cohort differences
4. The Opportunity — where to invest
5. Strategic Recommendations — actionable next steps with SO WHAT?

IMI CONTEXT: Founded 55+ years ago in Toronto. Managing Partner: Don Mayo. Offices: Toronto, Melbourne, Tokyo. Core methodology: Pulse™ surveys. Proprietary concept: The Say-Do Gap™."""
    if team_context:
        system_msg += f"\n\n{team_context}"

    # Inject survey data if a survey is loaded
    if survey_id and survey_id in _SURVEY_STORE:
        survey = _SURVEY_STORE[survey_id]
        if survey.get("structured_data"):
            data_inject = _json.dumps(survey["structured_data"], indent=2)
        else:
            data_inject = survey.get("raw_text", "")
        system_msg += f"\n\n<survey_data>\n{data_inject}\n</survey_data>\n\nThe above is the ONLY data you may reference. Every number in your response must come from this data."

    messages = [{"role": "system", "content": system_msg}]
    messages += [{"role": h.get("role", "user"), "content": h.get("content", "")} for h in history]
    messages.append({"role": "user", "content": message})

    # Adjust temperature based on query type
    temp = 0.1
    strategic_triggers = ['should', 'recommend', 'strategy', 'implication', 'opportunity']
    deck_triggers = ['deck', 'presentation', 'slides', 'full analysis', 'insight report', 'analyze']
    if any(t in message.lower() for t in deck_triggers):
        temp = 0.4
    elif any(t in message.lower() for t in strategic_triggers):
        temp = 0.3

    ollama_payload = {"model": "klaus-imi", "messages": messages, "stream": True, "options": {"temperature": temp, "top_p": 0.85, "num_ctx": 32768}}

    req = _ollama_client.build_request("POST", f"{OLLAMA_BASE}/api/chat",
                                        content=_json.dumps(ollama_payload).encode(),
                                        headers={"content-type": "application/json"})
    resp = await _ollama_client.send(req, stream=True)
    async def stream_imi():
        try:
            async for line in resp.aiter_lines():
                yield line.encode() + b"\n"
        finally:
            await resp.aclose()
    return StreamingResponse(stream_imi(), status_code=resp.status_code,
                             media_type="application/x-ndjson")


# ============================================
# SURVEY ENGINE — IMI-Grade Analysis
# ============================================

import pandas as pd
import io as _io
import tempfile as _tempfile

# IMI Deck Builder
import sys as _sys
_sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from survey_bridge import survey_to_slides
from deck_renderer import build_deck
from flat_adapter import flat_to_structured
from narrative_polisher import polish_configs

# _SURVEY_STORE declared near app init (line ~111)
_SURVEY_DIR = os.path.expanduser("~/.axe/memory/surveys")
os.makedirs(_SURVEY_DIR, exist_ok=True)


def _parse_survey_file(filename: str, content: bytes) -> tuple[pd.DataFrame, dict]:
    """Parse xlsx/csv/json into DataFrame + metadata."""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    if ext in ("xlsx", "xls"):
        xls = pd.ExcelFile(_io.BytesIO(content))
        # Find the sheet with actual data (most rows)
        best_sheet = None
        best_rows = 0
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
            df = df.dropna(how="all")
            if len(df) > best_rows:
                best_rows = len(df)
                best_sheet = sheet_name
        df = pd.read_excel(xls, sheet_name=best_sheet, header=None)
    elif ext == "csv":
        df = pd.read_csv(_io.BytesIO(content), header=None)
    elif ext == "json":
        df = pd.read_json(_io.BytesIO(content))
    else:
        df = pd.read_csv(_io.BytesIO(content), header=None, sep="\t")

    # Clean: drop fully empty rows and columns
    df = df.dropna(how="all").dropna(axis=1, how="all")

    metadata = {
        "filename": filename,
        "rows": len(df),
        "columns": len(df.columns),
        "sheet": best_sheet if ext in ("xlsx", "xls") else None,
    }

    return df, metadata


def _df_to_text(df: pd.DataFrame, max_rows: int = 100) -> str:
    """Convert DataFrame to readable text for LLM consumption."""
    # Use full precision for the data
    with pd.option_context('display.max_columns', None, 'display.width', None,
                           'display.max_colwidth', 50, 'display.float_format', lambda x: f'{x:.4f}' if abs(x) < 1 else f'{x:.2f}'):
        subset = df.head(max_rows)
        return subset.to_string(index=False)


def _parse_crosstab_structured(file_bytes: bytes, filename: str) -> dict | None:
    """
    Smart crosstab parser: detects segment headers + question blocks.
    Returns structured JSON for <survey_data> injection, or None if detection fails.
    Falls back to raw text mode if structure can't be detected.
    """
    try:
        df = pd.read_excel(_io.BytesIO(file_bytes), header=None)
    except Exception:
        return None

    # --- Detect segment header row ---
    segment_keywords = [
        'canada', 'total', 'female', 'male', 'west', 'ontario',
        'quebec', 'atlantic', '18 to', '35 to', '55 and',
        'born in', 'born outside', 'under $', '$50', '$125',
        'national', 'overall', 'age', 'gender', 'region', 'income'
    ]

    segment_row = None
    for idx, row in df.iterrows():
        row_text = ' '.join(str(v).lower() for v in row.values if pd.notna(v))
        matches = sum(1 for kw in segment_keywords if kw in row_text)
        if matches >= 3:
            segment_row = idx
            break

    if segment_row is None:
        return None

    # --- Build segment mapping ---
    segments = {}
    for col_idx, val in enumerate(df.iloc[segment_row]):
        if pd.notna(val) and str(val).strip():
            segments[col_idx] = str(val).strip()

    # --- Detect question blocks ---
    questions = []
    current_q = None
    last_base_sizes = {}

    for idx in range(segment_row + 1, len(df)):
        row = df.iloc[idx]
        first_cell = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''

        # Question header (Q followed by number)
        if first_cell and first_cell.startswith('Q') and any(c.isdigit() for c in first_cell[:4]):
            if current_q:
                questions.append(current_q)
            current_q = {
                'id': first_cell.split(' ')[0] if ' ' in first_cell else first_cell,
                'text': first_cell,
                'options': []
            }
            continue

        # Sample size row
        if 'sample' in first_cell.lower() or 'n=' in first_cell.lower() or 'base' in first_cell.lower():
            base_sizes = {}
            for col_idx, seg_label in segments.items():
                val = row.iloc[col_idx] if col_idx < len(row) else None
                if pd.notna(val):
                    try:
                        base_sizes[seg_label] = int(float(val))
                    except (ValueError, TypeError):
                        pass
            if current_q:
                current_q['base_sizes'] = base_sizes
            last_base_sizes = base_sizes
            continue

        # Option row (label + numeric values)
        if first_cell and current_q:
            values = {}
            has_numeric = False
            for col_idx, seg_label in segments.items():
                val = row.iloc[col_idx] if col_idx < len(row) else None
                if pd.notna(val):
                    try:
                        num_val = float(val)
                        if 0 < num_val < 1:
                            num_val = round(num_val * 100, 1)
                        values[seg_label] = num_val
                        has_numeric = True
                    except (ValueError, TypeError):
                        pass

            if has_numeric:
                current_q['options'].append({
                    'label': first_cell,
                    'values': values
                })

    if current_q:
        questions.append(current_q)

    if not questions:
        return None

    # Determine total_n from last known base sizes
    total_n = 0
    if last_base_sizes:
        for key in ['Canada', 'Total', 'National', 'Overall']:
            if key in last_base_sizes:
                total_n = last_base_sizes[key]
                break
        if total_n == 0 and last_base_sizes:
            total_n = list(last_base_sizes.values())[0]

    return {
        'survey_name': filename.replace('.xlsx', '').replace('.csv', '').replace('_', ' '),
        'total_n': total_n,
        'segments': list(segments.values()),
        'questions': questions
    }


def _validate_response(response_text: str, survey_data: dict) -> dict:
    """
    Post-generation anti-hallucination check.
    Scans response for percentages, cross-references against source data.
    """
    warnings = []

    # Extract all percentages from response
    percentages = _re.findall(r'(\d+(?:\.\d+)?)\s*%', response_text)

    # Build set of valid numbers from survey data
    valid_numbers = set()
    for q in survey_data.get('questions', []):
        for option in q.get('options', []):
            for seg, val in option.get('values', {}).items():
                valid_numbers.add(str(round(val, 1)))
                valid_numbers.add(str(int(val)))
        for seg, val in q.get('base_sizes', {}).items():
            valid_numbers.add(str(val))

    for pct in percentages:
        if pct not in valid_numbers:
            try:
                pct_float = float(pct)
                # Allow derived metrics (gaps, indices, combined scores)
                if pct_float > 100:
                    continue
            except ValueError:
                pass
            warnings.append(f"{pct}% not found in source data")

    # Detect hedging language before numbers (hallucination signal)
    hedge_patterns = [
        r'approximately\s+\d', r'roughly\s+\d', r'about\s+\d',
        r'I believe\s+\d', r'typically\s+\d', r'usually\s+\d'
    ]
    for pattern in hedge_patterns:
        if _re.search(pattern, response_text, _re.IGNORECASE):
            warnings.append("Hedging language before number detected")
            break

    return {"valid": len(warnings) == 0, "warnings": warnings[:10]}


def _get_temperature(question: str) -> float:
    """Dynamic temperature based on query type."""
    q_lower = question.lower()
    deck_triggers = ['deck', 'presentation', 'slides', 'full analysis', 'insight report', 'analyze']
    strategic_triggers = ['should', 'recommend', 'strategy', 'implication', 'opportunity']

    if any(t in q_lower for t in deck_triggers):
        return 0.4
    elif any(t in q_lower for t in strategic_triggers):
        return 0.3
    else:
        return 0.1


def _build_imi_analysis_prompt(survey_text: str, filename: str) -> str:
    """Build the IMI-style analysis prompt."""
    return f"""You are an elite data analyst at IMI International / ConsultIMI — a global consultancy known for "Insight. Driving. Profit."

You have been given raw survey crosstab data from the file "{filename}". Your job is to produce a full IMI-grade insight report.

SURVEY DATA:
```
{survey_text}
```

Produce a comprehensive analysis following this EXACT structure. Use the actual numbers from the data above — never fabricate statistics.

## REPORT STRUCTURE:

### 1. EXECUTIVE SUMMARY
- 3 big-stat callout cards (the most surprising/important numbers)
- Each stat should have a bold number and a 1-sentence insight

### 2. NATIONAL RESULTS (Overall/Total column)
- Rank all response options from highest to lowest
- Show percentages
- Note what's surprising vs expected

### 3. COMBINED ANALYSIS (if 1st + 2nd choice or combined data exists)
- Show combined reach/preference
- Identify the real competitive landscape
- Note any virtual ties or surprising rankings

### 4. DEMOGRAPHIC DEEP-DIVES
For EACH demographic break in the crosstab (age, gender, region, income, nativity, etc.):
- Identify the biggest over-index and under-index vs national average
- Calculate index scores (demo % / national % × 100)
- Highlight statistically meaningful gaps (>5 percentage points)
- Name the key insight for each demographic

### 5. THE HEADLINE INSIGHT
- What's the single most important finding?
- The "showstopper" slide — the insight that would make a CMO sit up
- Include the specific comparison numbers

### 6. STRATEGIC IMPLICATIONS
- 4 actionable "so what?" recommendations
- Each should name: the insight, the audience, and the sponsor/brand action
- Frame as business opportunities, not just data observations

### 7. KEY QUOTES FOR SLIDES
- For each major finding, write a "KEY INSIGHT" or "SO WHAT?" callout (1-2 sentences, punchy, executive-ready)

FORMATTING RULES:
- Use actual percentages from the data (convert decimals to % where needed — e.g., 0.27 = 27%)
- Calculate index scores where relevant (demo/total × 100)
- Bold the most important numbers
- Use IMI's signature phrases: "Insight. Driving. Profit.", "KEY INSIGHT:", "SO WHAT?"
- Be specific — cite sample sizes, exact percentages, exact gaps
- If multiple questions exist in the data, analyze each one
- End with a note on methodology (sample size, crosstab dimensions)

Write the full report now."""


@app.post("/klaus/survey/load")
async def survey_load(request: Request):
    """Load a survey file (xlsx, csv, json). Accepts multipart form or JSON with base64."""
    content_type = request.headers.get("content-type", "")

    if "multipart" in content_type:
        form = await request.form()
        file = form.get("file")
        if not file:
            return JSONResponse({"error": "No file provided"}, status_code=400)
        filename = file.filename
        content = await file.read()
    else:
        body = await request.json()
        filename = body.get("filename", "upload.csv")
        import base64
        content = base64.b64decode(body.get("data", ""))

    try:
        df, metadata = _parse_survey_file(filename, content)
    except Exception as e:
        return JSONResponse({"error": f"Failed to parse file: {str(e)}"}, status_code=400)

    survey_id = str(uuid.uuid4())[:8]
    raw_text = _df_to_text(df, max_rows=200)

    # Try smart structured parsing: crosstab first, then flat format
    structured = _parse_crosstab_structured(content, filename)
    if not structured:
        structured = flat_to_structured(content, filename)

    _SURVEY_STORE[survey_id] = {
        "name": filename,
        "df": df,
        "raw_text": raw_text,
        "metadata": metadata,
        "structured_data": structured,
    }

    # Persist raw text for future sessions
    with open(os.path.join(_SURVEY_DIR, f"{survey_id}.txt"), "w") as f:
        f.write(f"# {filename}\n\n{raw_text}")

    return JSONResponse({
        "survey_id": survey_id,
        "name": filename,
        "columns": list(df.columns) if not df.empty else [],
        "row_count": len(df),
        "preview": raw_text[:1000],
    })


@app.get("/klaus/survey/list")
async def survey_list():
    """List loaded surveys."""
    surveys = []
    for sid, s in _SURVEY_STORE.items():
        surveys.append({
            "id": sid,
            "name": s["name"],
            "rows": s["metadata"]["rows"],
            "columns": s["metadata"]["columns"],
        })
    return JSONResponse({"surveys": surveys})


@app.get("/klaus/survey/{survey_id}/summary")
async def survey_summary(survey_id: str):
    """Get survey summary."""
    survey = _SURVEY_STORE.get(survey_id)
    if not survey:
        return JSONResponse({"error": "Survey not found"}, status_code=404)

    df = survey["df"]
    return JSONResponse({
        "survey_id": survey_id,
        "name": survey["name"],
        "total_rows": len(df),
        "total_columns": len(df.columns),
        "preview": survey["raw_text"][:2000],
        "metadata": survey["metadata"],
    })


@app.post("/klaus/survey/query")
async def survey_query(request: Request):
    """Query a survey with natural language — sends data + question to Ollama."""
    body = await request.json()
    survey_id = body.get("survey_id", "")
    question = body.get("question", "")

    survey = _SURVEY_STORE.get(survey_id)
    if not survey:
        return JSONResponse({"error": "Survey not found"}, status_code=404)

    # Use structured data if available, fall back to raw text
    structured = survey.get("structured_data")
    if structured:
        data_block = f"<survey_data>\n{_json.dumps(structured, indent=2)}\n</survey_data>\n\nThe above is the ONLY data you may reference. Every number in your response must come from this data."
    else:
        data_block = f"<survey_data>\nSURVEY: {survey['name']}\n{survey['raw_text']}\n</survey_data>\n\nThe above is the ONLY data you may reference. Every number in your response must come from this data."

    messages = [
        {"role": "system", "content": data_block},
        {"role": "user", "content": question},
    ]

    temperature = _get_temperature(question)

    ollama_payload = {
        "model": "klaus-imi",
        "messages": messages,
        "stream": True,
        "options": {"temperature": temperature, "top_p": 0.85, "num_ctx": 32768},
    }

    req = _ollama_client.build_request(
        "POST", f"{OLLAMA_BASE}/api/chat",
        content=_json.dumps(ollama_payload).encode(),
        headers={"content-type": "application/json"},
    )
    resp = await _ollama_client.send(req, stream=True)

    async def stream_response():
        try:
            async for line in resp.aiter_lines():
                yield line.encode() + b"\n"
        finally:
            await resp.aclose()

    return StreamingResponse(stream_response(), status_code=resp.status_code,
                             media_type="application/x-ndjson")


@app.post("/klaus/survey/{survey_id}/analyze")
async def survey_analyze(survey_id: str):
    """Full IMI-style analysis — the 11-slide framework as streaming response."""
    survey = _SURVEY_STORE.get(survey_id)
    if not survey:
        return JSONResponse({"error": "Survey not found"}, status_code=404)

    prompt = _build_imi_analysis_prompt(survey["raw_text"], survey["name"])

    # Use structured data injection if available
    structured = survey.get("structured_data")
    if structured:
        data_inject = f"<survey_data>\n{_json.dumps(structured, indent=2)}\n</survey_data>"
    else:
        data_inject = f"<survey_data>\n{survey['raw_text']}\n</survey_data>"

    messages = [
        {"role": "system", "content": f"{data_inject}\n\nThe above is the ONLY data you may reference."},
        {"role": "user", "content": prompt},
    ]

    ollama_payload = {
        "model": "klaus-imi",
        "messages": messages,
        "stream": True,
        "options": {"temperature": 0.4, "top_p": 0.85, "num_ctx": 32768},
    }

    req = _ollama_client.build_request(
        "POST", f"{OLLAMA_BASE}/api/chat",
        content=_json.dumps(ollama_payload).encode(),
        headers={"content-type": "application/json"},
    )
    resp = await _ollama_client.send(req, stream=True)

    async def stream_analysis():
        try:
            async for line in resp.aiter_lines():
                yield line.encode() + b"\n"
        finally:
            await resp.aclose()

    return StreamingResponse(stream_analysis(), status_code=resp.status_code,
                             media_type="application/x-ndjson")


@app.post("/klaus/survey/{survey_id}/deck")
async def survey_deck(survey_id: str):
    """Generate IMI-branded .pptx deck from loaded survey data."""
    survey = _SURVEY_STORE.get(survey_id)
    if not survey:
        return JSONResponse({"error": "Survey not found"}, status_code=404)

    structured = survey.get("structured_data")
    if not structured:
        return JSONResponse(
            {"error": "No structured data available for this survey. Re-upload as xlsx crosstab."},
            status_code=400,
        )

    try:
        configs = survey_to_slides(structured)
        configs = polish_configs(configs, structured)
        output_path = os.path.join(_tempfile.gettempdir(), f"{survey_id}_deck.pptx")
        build_deck(configs, output_path)
    except Exception as e:
        return JSONResponse({"error": f"Deck generation failed: {str(e)}"}, status_code=500)

    safe_name = structured.get("survey_name", survey_id).replace(" ", "_")
    return FileResponse(
        output_path,
        media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
        filename=f"{safe_name}_IMI_Deck.pptx",
    )


@app.post("/klaus/survey/{survey_id}/chat-analysis")
async def survey_chat_analysis(survey_id: str):
    """
    Generate the full 11-section IMI analysis as structured text for chat display,
    AND pre-build the deck so a download link is ready immediately.
    Returns JSON with 'analysis' (markdown text) and 'deck_url'.
    """
    survey = _SURVEY_STORE.get(survey_id)
    if not survey:
        return JSONResponse({"error": "Survey not found"}, status_code=404)

    structured = survey.get("structured_data")
    if not structured:
        return JSONResponse({"error": "No structured data for this survey."}, status_code=400)

    try:
        configs = survey_to_slides(structured)
        configs = polish_configs(configs, structured)
    except Exception as e:
        return JSONResponse({"error": f"Analysis failed: {str(e)}"}, status_code=500)

    # Build human-readable markdown from the slide configs
    analysis = _configs_to_markdown(configs, structured)

    # Pre-build the deck
    deck_path = os.path.join(_tempfile.gettempdir(), f"{survey_id}_deck.pptx")
    try:
        build_deck(configs, deck_path)
        deck_ready = True
    except Exception:
        deck_ready = False

    return JSONResponse({
        "survey_id": survey_id,
        "survey_name": structured.get("survey_name", ""),
        "analysis": analysis,
        "deck_available": deck_ready,
        "deck_url": f"/klaus/survey/{survey_id}/deck" if deck_ready else None,
    })


def _configs_to_markdown(configs: dict, structured: dict) -> str:
    """Convert slide configs into a formatted markdown analysis for chat."""
    lines = []
    name = structured.get("survey_name", "Survey")
    total_n = structured.get("total_n", 0)

    # 1. Title
    s1 = configs.get("slide_1", {})
    lines.append(f"# {s1.get('title', name)}")
    lines.append(f"*{s1.get('subtitle', '')}*")
    lines.append(f"_{s1.get('methodology', f'n = {total_n}')}_")
    lines.append("")

    # 2. Executive Summary
    s2 = configs.get("slide_2", {})
    lines.append("## Executive Summary")
    for stat in s2.get("stats", []):
        lines.append(f"- **{stat['number']}** — {stat['description']}")
    lines.append("")

    # 3. National Results
    s3 = configs.get("slide_3", {})
    lines.append(f"## {s3.get('title', 'National Results')}")
    if s3.get("subtitle"):
        lines.append(f"*{s3['subtitle']}*")
    for label, val in s3.get("data", {}).items():
        lines.append(f"- **{val}%** — {label}")
    if s3.get("insight"):
        lines.append(f"\n> {s3['insight']}")
    lines.append("")

    # 4. Combined Preference
    s4 = configs.get("slide_4", {})
    lines.append(f"## {s4.get('title', 'Combined Preference')}")
    if s4.get("subtitle"):
        lines.append(f"*{s4['subtitle']}*")
    for label, val in s4.get("data", {}).items():
        pct = f"{val}%" if isinstance(val, (int, float)) else str(val)
        lines.append(f"- **{pct}** — {label}")
    if s4.get("insight"):
        lines.append(f"\n> {s4['insight']}")
    lines.append("")

    # 5. Biggest Divide
    s5 = configs.get("slide_5", {})
    lines.append(f"## {s5.get('title', 'The Biggest Divide')}")
    if s5.get("subtitle"):
        lines.append(f"*{s5['subtitle']}*")
    left_label = s5.get("left_label", "Group A")
    right_label = s5.get("right_label", "Group B")
    left_n = f" (n={s5['left_n']})" if s5.get("left_n") else ""
    right_n = f" (n={s5['right_n']})" if s5.get("right_n") else ""
    lines.append(f"\n**{left_label}**{left_n}:")
    for label, val in s5.get("left_data", {}).items():
        lines.append(f"  - {val}% — {label}")
    lines.append(f"\n**{right_label}**{right_n}:")
    for label, val in s5.get("right_data", {}).items():
        lines.append(f"  - {val}% — {label}")
    if s5.get("so_what"):
        lines.append(f"\n**SO WHAT?** {s5['so_what']}")
    lines.append("")

    # 6. Age/Generational
    s6 = configs.get("slide_6", {})
    lines.append(f"## {s6.get('title', 'Generational Lens')}")
    for label, age_vals in s6.get("data", {}).items():
        if isinstance(age_vals, dict):
            parts = [f"{k}: {v}%" for k, v in age_vals.items()]
            lines.append(f"- **{label}**: {' | '.join(parts)}")
    if s6.get("insight"):
        lines.append(f"\n> {s6['insight']}")
    lines.append("")

    # 7. Gender & Regional
    s7 = configs.get("slide_7", {})
    lines.append(f"## {s7.get('title', 'Gender & Regional Lens')}")
    if s7.get("gender_data"):
        lines.append("\n| Event | Female | Male | Gap |")
        lines.append("|-------|--------|------|-----|")
        for row in s7["gender_data"]:
            lines.append(f"| {' | '.join(str(c) for c in row)} |")
    if s7.get("regions"):
        lines.append("\n**Regional highlights:**")
        for r in s7["regions"]:
            lines.append(f"- **{r['name']}**: {r['insight']}")
    if s7.get("so_what"):
        lines.append(f"\n**SO WHAT?** {s7['so_what']}")
    lines.append("")

    # 8. Aspirational
    s8 = configs.get("slide_8", {})
    lines.append(f"## {s8.get('title', 'Preference Landscape')}")
    for label, val in s8.get("data", {}).items():
        lines.append(f"- **{val}%** — {label}")
    if s8.get("insight"):
        lines.append(f"\n> {s8['insight']}")
    lines.append("")

    # 9. Income
    s9 = configs.get("slide_9", {})
    lines.append(f"## {s9.get('title', 'Income Analysis')}")
    for label, inc_vals in s9.get("data", {}).items():
        if isinstance(inc_vals, dict):
            parts = [f"{k}: {v}%" for k, v in inc_vals.items()]
            lines.append(f"- **{label}**: {' | '.join(parts)}")
    if s9.get("insight"):
        lines.append(f"\n> {s9['insight']}")
    lines.append("")

    # 10. Strategic Implications
    s10 = configs.get("slide_10", {})
    lines.append("## Strategic Implications")
    for i, rec in enumerate(s10.get("recommendations", []), 1):
        lines.append(f"\n### {i}. {rec['title']}")
        lines.append(rec["body"])
    lines.append("")

    # 11. Methodology
    lines.append("---")
    lines.append(f"*Source: IMI Pulse™ | {name} | n = {total_n}*")

    return "\n".join(lines)


@app.post("/klaus/survey/upload-and-analyze")
async def survey_upload_and_analyze(request: Request):
    """One-shot: upload file and get full IMI analysis. Accepts multipart form."""
    content_type = request.headers.get("content-type", "")

    if "multipart" not in content_type:
        return JSONResponse({"error": "Must use multipart/form-data"}, status_code=400)

    form = await request.form()
    file = form.get("file")
    if not file:
        return JSONResponse({"error": "No file provided"}, status_code=400)

    filename = file.filename
    content = await file.read()

    try:
        df, metadata = _parse_survey_file(filename, content)
    except Exception as e:
        return JSONResponse({"error": f"Failed to parse: {str(e)}"}, status_code=400)

    survey_id = str(uuid.uuid4())[:8]
    raw_text = _df_to_text(df, max_rows=200)

    # Try smart structured parsing: crosstab first, then flat format
    structured = _parse_crosstab_structured(content, filename)
    if not structured:
        structured = flat_to_structured(content, filename)

    _SURVEY_STORE[survey_id] = {
        "name": filename,
        "df": df,
        "raw_text": raw_text,
        "metadata": metadata,
        "structured_data": structured,
    }

    prompt = _build_imi_analysis_prompt(raw_text, filename)

    # Use structured data if available
    if structured:
        data_inject = f"<survey_data>\n{_json.dumps(structured, indent=2)}\n</survey_data>"
    else:
        data_inject = f"<survey_data>\n{raw_text}\n</survey_data>"

    messages = [
        {"role": "system", "content": f"{data_inject}\n\nThe above is the ONLY data you may reference."},
        {"role": "user", "content": prompt},
    ]

    ollama_payload = {
        "model": "klaus-imi",
        "messages": messages,
        "stream": True,
        "options": {"temperature": 0.4, "top_p": 0.85, "num_ctx": 32768},
    }

    req = _ollama_client.build_request(
        "POST", f"{OLLAMA_BASE}/api/chat",
        content=_json.dumps(ollama_payload).encode(),
        headers={"content-type": "application/json"},
    )
    resp = await _ollama_client.send(req, stream=True)

    async def stream_full():
        # First emit metadata
        yield _json.dumps({"type": "metadata", "survey_id": survey_id, "name": filename, "rows": len(df), "columns": len(df.columns)}).encode() + b"\n"
        # Then stream analysis
        try:
            async for line in resp.aiter_lines():
                yield line.encode() + b"\n"
        finally:
            await resp.aclose()

    return StreamingResponse(stream_full(), status_code=200,
                             media_type="application/x-ndjson")


@app.post("/klaus/file/upload")
async def klaus_file_upload(request: Request):
    """General file upload for chat context. Parses xlsx/csv and returns summary."""
    content_type = request.headers.get("content-type", "")
    if "multipart" not in content_type:
        return JSONResponse({"success": False, "error": "Must use multipart/form-data"}, status_code=400)

    form = await request.form()
    file = form.get("file")
    if not file:
        return JSONResponse({"success": False, "error": "No file provided"}, status_code=400)

    filename = file.filename
    content = await file.read()

    try:
        df, metadata = _parse_survey_file(filename, content)
    except Exception as e:
        return JSONResponse({"success": False, "error": f"Failed to parse file: {str(e)}", "filename": filename})

    upload_id = str(uuid.uuid4())[:8]
    raw_text = _df_to_text(df, max_rows=50)

    # Store for later reference
    _SURVEY_STORE[upload_id] = {
        "name": filename,
        "df": df,
        "raw_text": raw_text,
        "metadata": metadata,
        "structured_data": None,
    }

    columns = [str(c) for c in df.columns.tolist()]
    summary = f"**{filename}** — {len(df)} rows, {len(columns)} columns\n\nColumns: {', '.join(columns[:15])}"
    if len(columns) > 15:
        summary += f" (+{len(columns)-15} more)"

    return JSONResponse({
        "success": True,
        "filename": filename,
        "upload_id": upload_id,
        "summary": summary,
        "columns": columns,
        "row_count": len(df),
    })


# ============================================
# CORBOT — Agentic coding assistant
# ============================================

from corbot_tools import TOOL_DEFINITIONS, execute_tool

CORBOT_SYSTEM = """You are Corbot, an agentic coding assistant with direct access to the filesystem and shell. You were built by James Lewis as part of the AXE platform.

You have the following tools available — use them proactively to answer questions and complete tasks:

TOOLS:
- read_file(path) — Read any file's contents
- write_file(path, content) — Create or overwrite a file
- edit_file(path, old_text, new_text) — Find and replace text in a file (precise string match)
- list_directory(path) — List files and folders with sizes and types
- run_command(command, working_dir?) — Execute any shell command (bash). 30s timeout.
- search_files(pattern, path?) — Find files by glob pattern (e.g. "**/*.py", "src/**/*.tsx")
- search_content(regex, path?) — Search file contents with regex (like grep/ripgrep)

RULES:
1. Always use tools to verify — never guess file contents or command output.
2. When editing files, read them first to get exact text for the old_text parameter.
3. For multi-step tasks, chain tools: read → understand → edit/write → verify.
4. Be direct and concise. Lead with results, not process narration.
5. When asked about your capabilities, list these actual tools and what they do.
6. Show file paths and command outputs — the user sees tool executions inline.

You operate on a Mac Studio running macOS. The backend is FastAPI + Ollama (Qwen 2.5)."""


@app.post("/corbot/chat")
async def corbot_chat(request: Request):
    """Agentic coding chat — Qwen + tool loop, streamed via SSE."""
    body = await request.json()
    messages = body.get("messages", [])
    model = body.get("model", "qwen2.5:32b")
    working_dir = body.get("working_dir", os.path.expanduser("~"))
    max_iterations = 10  # Safety limit

    # Prepend system prompt
    full_messages = [{"role": "system", "content": CORBOT_SYSTEM}] + messages

    async def agentic_stream():
        nonlocal full_messages
        iterations = 0

        while iterations < max_iterations:
            iterations += 1

            # Call Ollama with tools
            ollama_payload = {
                "model": model,
                "messages": full_messages,
                "tools": TOOL_DEFINITIONS,
                "stream": False,
                "options": {"temperature": 0.1, "num_ctx": 32768}
            }

            try:
                resp = await _ollama_client.post(
                    f"{OLLAMA_BASE}/api/chat",
                    json=ollama_payload,
                    headers={"content-type": "application/json"}
                )
                if resp.status_code != 200:
                    yield _json.dumps({"type": "error", "content": f"Ollama error: {resp.status_code}"}).encode() + b"\n"
                    break

                result = resp.json()
                msg = result.get("message", {})
                content = msg.get("content", "")
                tool_calls = msg.get("tool_calls", [])

            except Exception as e:
                yield _json.dumps({"type": "error", "content": f"Ollama connection error: {str(e)}"}).encode() + b"\n"
                break

            # If no tool calls, we're done — stream the final text
            if not tool_calls:
                if content:
                    yield _json.dumps({"type": "text", "content": content}).encode() + b"\n"
                yield _json.dumps({"type": "done"}).encode() + b"\n"
                break

            # If there's text before tool calls, stream it
            if content:
                yield _json.dumps({"type": "text", "content": content}).encode() + b"\n"

            # Execute each tool call
            full_messages.append(msg)  # Add assistant message with tool_calls

            for tc in tool_calls:
                fn = tc.get("function", {})
                tool_name = fn.get("name", "")
                tool_args = fn.get("arguments", {})

                # Stream tool call event
                yield _json.dumps({"type": "tool_call", "name": tool_name, "args": tool_args}).encode() + b"\n"

                # Execute
                tool_result = execute_tool(tool_name, tool_args, working_dir)

                # Stream tool result
                yield _json.dumps({"type": "tool_result", "name": tool_name, **tool_result}).encode() + b"\n"

                # Add tool result to messages for next iteration
                full_messages.append({
                    "role": "tool",
                    "content": tool_result.get("output", "")
                })

        else:
            yield _json.dumps({"type": "error", "content": "Max tool iterations reached"}).encode() + b"\n"
            yield _json.dumps({"type": "done"}).encode() + b"\n"

    return StreamingResponse(agentic_stream(), status_code=200,
                             media_type="application/x-ndjson")


# ============================================
# MAIN
# ============================================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
